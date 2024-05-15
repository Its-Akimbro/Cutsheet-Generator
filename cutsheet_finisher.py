import csv
import openpyxl as xl
import os

legacy_switch_name = ''
new_switch_name = ''
room_number = ''
legacy_ip = ''
new_switch_name_list = []
wap_vlan = ''

total_ports = 49

class new_switch():
    def __init__(self, switch_model: str,filled_blades: dict):
        self.switch_model = switch_model
        self.filled_blades = filled_blades

def legacy_switch_name_finder(file_name):
    global legacy_switch_name

    with open(file_name) as pre_mapped_file:
        for name in pre_mapped_file:
            list_of_name = name.split(',') 
            if 'Legacy Switch' not in list_of_name[0] and len(list_of_name) > 5:
                legacy_switch_name = list_of_name[0]
                break

def new_switch_name_finder(file_name):
    global new_switch_name

    with open(file_name) as pre_mapped_file:
        for name in pre_mapped_file:
            list_of_name = name.split(',')
            if 'Legacy Switch' not in list_of_name[0] and len(list_of_name) > 5:
                new_switch_name = list_of_name[4]
                break

def room_number_finder(file_name):
    global room_number

    with open(file_name) as pre_mapped_file:
        for name in pre_mapped_file:
            list_of_name = name.split(',')
            if 'Legacy Switch' not in list_of_name[0] and len(list_of_name) > 5:
                list_of_switch_name = list_of_name[4].split('-')
                room_number = list_of_switch_name[2]
                break

def legacy_ip_finder(file_name):
    global legacy_ip

    with open(file_name) as pre_mapped_file:
        for name in pre_mapped_file:
            list_of_name = name.split(',')
            if 'Legacy Switch IP' not in list_of_name[0] and len(list_of_name) > 5:
                legacy_ip = list_of_name[1]
                break

def max_port_finder(wap_file, switch_count):
    with open(wap_file) as wireless_ap_count:
        global room_number

        ap_count = 0

        for aps in wireless_ap_count:

            list_of_aps = aps.split(",")

            if room_number in list_of_aps[3] and room_number in list_of_aps[5]:
                ap_count += 1

        if switch_count == 1:
            total_ports = 96 - ap_count
            return round(total_ports/2)-3
        
        elif switch_count > 1:
            total_ports = (96*switch_count) - ap_count
            return round(total_ports/(2*switch_count))-3

def port_mapping_file_info_to_cutsheet(file_name, patch_panel_file,switch_count, filled_blade):
    with open(patch_panel_file, newline='') as port_mapped:
        with open(file_name, newline='') as pre_portmapped:
            with open(f'{file_name[:file_name.find('.')]} + patch_panel.csv', 'w', newline='') as pre_portmapped_and_patch_panel_info:

                row = csv.writer(pre_portmapped_and_patch_panel_info)
                
                count = 0

                for interfaces in enumerate(pre_portmapped):
                    list_of_interfaces = interfaces[1].strip('\r\n').split(',')

                    if 0 == interfaces[0]:  
                        row.writerow(list_of_interfaces)
                        continue

                    for patch_panel_and_switch_info in port_mapped:
                        list_of_patch_panel_and_switch_info = patch_panel_and_switch_info.split(',')
   
                        if list_of_interfaces[2] == list_of_patch_panel_and_switch_info[2] and list_of_interfaces[0] == list_of_patch_panel_and_switch_info[0]:
                            if list_of_patch_panel_and_switch_info[3] != '' and list_of_patch_panel_and_switch_info[3].lower() != 'n/a' and list_of_patch_panel_and_switch_info[3].lower() != 'x':  
                                list_of_interfaces.insert(3,list_of_patch_panel_and_switch_info[3])  
                                list_of_interfaces.pop(4)
                            else:
                                continue
                    
                            # if 'blade ' in patch_panel_and_switch_info.lower():
                            #     blade_num = patch_panel_and_switch_info[patch_panel_and_switch_info.lower().index('blade') + 6]
                            #     list_of_interfaces.insert(10,f'Blade {blade_num}')  
                            #     list_of_interfaces.pop(11)

                            if list_of_patch_panel_and_switch_info[10].isnumeric() == True:
                                blade_num = list_of_patch_panel_and_switch_info[10]
                                list_of_interfaces.insert(10,f'Blade {blade_num}')  
                                list_of_interfaces.pop(11)

                            else:
                                list_of_interfaces.insert(10,f'Blade 1')  
                                list_of_interfaces.pop(11)
                            
                            if 'as-' in patch_panel_and_switch_info.lower():
                                switch_num = patch_panel_and_switch_info[patch_panel_and_switch_info.lower().index('as-') + 3]
                                if switch_num == '2':
                                    switch_2_namme = list_of_interfaces[4][:-1] + '2'
                                    list_of_interfaces.insert(4,switch_2_namme)  
                                    list_of_interfaces.pop(5)

                                    last_octet = int(list_of_interfaces[5][-1]) + 1
                                    switch_2_ip = list_of_interfaces[5][:-1] + str(last_octet)
                                    list_of_interfaces.insert(5,switch_2_ip)  
                                    list_of_interfaces.pop(6)

                                list_of_interfaces.insert(9,f'Switch {switch_num}')  
                                list_of_interfaces.pop(10)
                            else:
                                if switch_count > 1 and list_of_interfaces[9] == '':
                                    count += 1

                                    if count % 2 == True:
                                        calculated_switch_num = 1

                                    if count % 2 == False:
                                        calculated_switch_num = 2

                                    list_of_interfaces.insert(9,f'Switch {calculated_switch_num}')  
                                    list_of_interfaces.pop(10)
                                else:
                                    list_of_interfaces.insert(9,f'Switch 1') 
                                    list_of_interfaces.pop(10)

                            if 'a-half' in list_of_interfaces[13]:
                                list_of_filled_blades = filled_blade.split(',')
                                list_of_interfaces.insert(10,f'Blade {list_of_filled_blades[-1]}')  
                                list_of_interfaces.pop(11)

                            if 'a-10' == list_of_interfaces[12]:
                                list_of_filled_blades = filled_blade.split(',')
                                list_of_interfaces.insert(10,f'Blade {list_of_filled_blades[-1]}')  
                                list_of_interfaces.pop(11)
                                
                            row.writerow(list_of_interfaces)
                    port_mapped.seek(1)

def new_switch_port_adder(switch_details,max_port,file_name):
    global legacy_switch_name

    with open(f"{legacy_switch_name} cutsheet.csv", "w", newline='') as final_cutsheet:
        with open(f'{file_name[:file_name.find('.')]} + patch_panel.csv', newline='') as cutsheet_w_patchpanel_info:
            test = csv.writer(final_cutsheet)

            header = cutsheet_w_patchpanel_info.readline()
            final_cutsheet.write(header)

            for i in cutsheet_w_patchpanel_info:
                list_of_i = i.strip('\r\n').split(",")
                
                blade = int(list_of_i[10][6:])
                switch_num = int(list_of_i[9][7])

                if str(blade) not in str(switch_details[switch_num-1].filled_blades.keys()):
                    for i in switch_details[switch_num-1].filled_blades:
                        if int(i) > blade:
                            list_of_i[10] = f"blade {i}"
                            blade = i
                            break
                
                if blade == 1 or blade == 2:
                    if switch_details[switch_num-1].filled_blades[blade] == max_port:
                        for i in switch_details[switch_num-1].filled_blades:
                            if i > blade and switch_details[switch_num-1].filled_blades[i] != max_port:
                                list_of_i[10] = f"blade {i}"  
                                blade = i  
                                break

                if switch_details[switch_num-1].filled_blades[blade] == total_ports:
                    for i in switch_details[switch_num-1].filled_blades:
                        if i > blade and switch_details[switch_num-1].filled_blades[i] != total_ports:
                            list_of_i[10] = f"blade {i}"
                            blade = i
                            break

                if switch_details[switch_num-1].switch_model == "9300":
                    if 36 >= switch_details[switch_num-1].filled_blades[blade]:
                        list_of_i.insert(6,f"Tw{blade}/0/{switch_details[switch_num-1].filled_blades[blade]}")
                        list_of_i.pop(7)

                    elif 36 < switch_details[switch_num-1].filled_blades[blade]:
                        list_of_i.insert(6,f"Te{blade}/0/{switch_details[switch_num-1].filled_blades[blade]}")
                        list_of_i.pop(7)

                elif switch_details[switch_num-1].switch_model == "9407":
                    if 3 > blade:
                        list_of_i.insert(6,f"Fi{blade}/0/{switch_details[switch_num-1].filled_blades[blade]}")
                        list_of_i.pop(7)

                    else:
                        list_of_i.insert(6,f"Gi{blade}/0/{switch_details[switch_num-1].filled_blades[blade]}")
                        list_of_i.pop(7)

                elif switch_details[switch_num-1].switch_model == "9410":
                    if 3 > blade:
                        list_of_i.insert(6,f"Fi{blade}/0/{switch_details[switch_num-1].filled_blades[blade]}")
                        list_of_i.pop(7)
                        
                    else:
                        list_of_i.insert(6,f"Gi{blade}/0/{switch_details[switch_num-1].filled_blades[blade]}")
                        list_of_i.pop(7)

                switch_details[switch_num-1].filled_blades[blade] += 1
                test.writerow(list_of_i)

def list_of_switch_vlans(switch_details):
    global legacy_switch_name

    with open(f"{legacy_switch_name} cutsheet.csv") as finding_vlans:

        vlan_list = {i+1:[] for i in range(len(switch_details))}
        count = 0
        for i in finding_vlans:
            list_of_i = i.split(",")
            count += 1

            
            if 0 == count:  
                continue
            if 1 == count:  
                continue


            if list_of_i[4][int(-1)].isnumeric():
                switch_num = list_of_i[4][int(-1)]

                if list_of_i[14] not in vlan_list[int(switch_num)] and list_of_i[14].isnumeric():
                    vlan_list[int(switch_num)].append(list_of_i[14])

                if list_of_i[15] not in vlan_list[int(switch_num)] and list_of_i[15].isnumeric():
                    vlan_list[int(switch_num)].append(list_of_i[15])

        for i in vlan_list.keys():
            add_vlan_list = []
            for k in vlan_list[i]:
                add_vlan_list.append(f"vlan {k}")       
        
            vlan_list[i]=add_vlan_list

        return vlan_list

def multiple_new_switch_name_finder():
    global legacy_switch_name

    with open(f"{legacy_switch_name} cutsheet.csv") as cutsheet:
        switch_names = []

        for i in cutsheet:
            list_of_i = i.split(",")
            switch_name = list_of_i[4]

            if switch_name not in switch_names and switch_name != 'New Switch':
                if switch_name[-1] == '1':
                    switch_names.insert(0,switch_name)
                elif switch_name[-1] == '2':
                    switch_names.insert(1,switch_name)
                elif switch_name[-1] == '3':
                    switch_names.append(switch_name)

        return switch_names

def idf_config(site_vlans,switch_details):
    global legacy_switch_name
    global new_switch_name_list

    vlan_dict = list_of_switch_vlans(switch_details)
    new_switch_name_list = multiple_new_switch_name_finder()

    switch_list = [i+1 for i in range(len(switch_details))]

    vlan_dict_remoover = vlan_dict.copy()

    base_port_config = ['!\ndescription TYPE=user SW=END STATION IP=dhcp IF=NA\n!\n',
                        'switchport\n!\n',
                        'no logging event link-status\n!\n',
                        'no snmp trap link-status\n!\n',
                        'power inline auto\n!\n',
                        'storm-control broadcast level 1.00\n!\n',
                        'spanning-tree portfast\n!\n',
                        'spanning-tree bpduguard enable\n!\n',
                        'service-policy input L2-Access-Trust-In\n!\n',
                        'service-policy output L2-Access-Out\n!\n',
                        'ip dhcp snooping limit rate 30\n!\n',
                        'no shutdown\n!\n!\n',]

    for switch in range(len(switch_details)):
        switch_list[switch] = open(f"{new_switch_name_list} switchport config.txt","w")
        switch_list[switch].write(f"conf t\n!\n")

        with open(site_vlans) as switch_vlans:

            for i in switch_vlans:
                vlan = i[:i.find("\"")].lower()

                if vlan in vlan_dict[switch+1]:
                    with open(site_vlans) as vlan_name_finder:

                        for name in vlan_name_finder:
                            if name[:i.find("\"")].lower() == vlan and vlan in vlan_dict_remoover[switch+1]:
                                vlan_descr = vlan_name_finder.readline()
                                switch_list[switch].write(f"{vlan}\n{vlan_descr.lower()}!\n")
                                vlan_dict_remoover[switch+1].remove(vlan)
                        vlan_name_finder.close()
                        
        switch_vlans.close()
        
        with open(f"{legacy_switch_name} cutsheet.csv") as switchport_vlans:

            count = 0
            for switch_port_details in switchport_vlans:                
                list_of_switch_port_details = switch_port_details.split(',')
                
                try:
                    switch_num = int(list_of_switch_port_details[9][-1])
                except ValueError:
                    continue

                count += 1
   
                if switch_num-1 == switch:
                    access_vlan = list_of_switch_port_details[14]
                    voice_vlan = list_of_switch_port_details[15]
                    new_interface = list_of_switch_port_details[6]

                    if switch_details[switch].switch_model == '9407' or  switch_details[switch].switch_model == '9410':
                        if count == 1 or count == 2:
                            filled_blades = switch_details[switch].filled_blades

                            for blade in filled_blades.keys():
                                if blade == 1 or blade == 2:
                                    interface = f'interface range Fi{blade}/0/1-48\n'
                                    base_port_config.insert(0,interface)

                                else:
                                    interface = f'interface range gi{blade}/0/1-48\n'
                                    base_port_config.insert(0,interface)

                                switch_list[switch].write(''.join(base_port_config))
                                base_port_config.pop(0)

                    if switch_details[switch].switch_model == '9300' and count == 1:
                        filled_blades = switch_details[switch].filled_blades

                        for blade in filled_blades.keys():
                            interface = f'interface range TwoGigabitEthernet{blade}/0/1 - 36\n'
                            base_port_config.insert(0,interface)

                            switch_list[switch].write(''.join(base_port_config))
                            base_port_config.pop(0)

                            interface = f'interface range TenGigabitEthernet{blade}/0/37 - 48\n'
                            base_port_config.insert(0,interface)

                            switch_list[switch].write(''.join(base_port_config))
                            base_port_config.pop(0)

                    if voice_vlan != "":
                        switch_list[switch].write(f"interface {new_interface}\nswitchport access vlan {access_vlan}\nswitchport voice vlan {voice_vlan}\n!\n")
                        
                    elif voice_vlan == "":
                        switch_list[switch].write(f"interface {new_interface}\nswitchport access vlan {access_vlan}\n!\n")

            switchport_vlans.close()                                

def wap_vlan_finder(vlan_file):
    with open(vlan_file) as vlans:
        global room_number
        global wap_vlan

        vlan = ''
        description = ''

        for vlan_name in vlans:
            list_of_vlan_name = vlan_name.split()

            if list_of_vlan_name == []:
                continue

            if 'vlan' in list_of_vlan_name[0].lower():
                vlan = list_of_vlan_name[1]

            if 'name'in list_of_vlan_name[0].lower():
                description = list_of_vlan_name[1]

            if room_number.lower() in description.lower() and 'Legacy_' not in description:
                wap_vlan = vlan
                break

def wap_sheet(wap_file,switch_details):
    global legacy_switch_name
    global room_number
    global wap_vlan

    with open(wap_file) as wap_data:
        with open(f"wap cutsheet {legacy_switch_name}.csv","w") as wap_cutsheet:
            
            wap_cutsheet.write("PP Port,New Switch name,New Switch Port,New Patch Cord Color,Speed,Duplex,New Access Vlan,New Port Description,Mac Address,IP Address,Hostname,Serial Number,Asset Tag,Model\n")
        
            ap_swapper_count = {int(i+1):{1:[48,0],2:[48,0],3:[1,0]} for i in range(len(switch_details))}
            switch_count = 1

            for ap_info in wap_data:
                list_of_ap_info = ap_info.split(",")
            
                if room_number in list_of_ap_info[1] and room_number in list_of_ap_info[3]:

                    if ap_swapper_count[switch_count][3][1] == 2:
                        ap_swapper_count[switch_count][3][0] += 1

                    if ap_swapper_count[switch_count][3][1] == 4:
                        ap_swapper_count[switch_count][3][1] = 0
                        ap_swapper_count[switch_count][3][0] = 1
                        switch_count += 1

                        if switch_count > len(switch_details):
                            switch_count = 1
                    
                    new_switch_port = f"Fi{ap_swapper_count[switch_count][3][0]}/0/{ap_swapper_count[switch_count][ap_swapper_count[switch_count][3][0]][0]}"
                    ap_swapper_count[switch_count][3][1] += 1
                    ap_swapper_count[switch_count][ap_swapper_count[switch_count][3][0]][0] -= 1       

                    asset_tag = list_of_ap_info[8]
                    serial_number = list_of_ap_info[7]
                    mac_address = list_of_ap_info[6]
                    host_name = list_of_ap_info[2]
                    patch_panel_port = list_of_ap_info[10]    
                    site_code = list_of_ap_info[5]
                    building = list_of_ap_info[4]
                    room_num = list_of_ap_info[3]         
                    
                    formatted_ap_info = f'{patch_panel_port},{site_code}-{building}-{room_num}-as{switch_count},{new_switch_port},purple,a-5000,a-full,vlan {wap_vlan},AP={host_name} SN={serial_number} HW=C9130AXI-B,{mac_address},DHCP,{host_name},{serial_number},{asset_tag},C9130AXI-B\n'
                    
                    wap_cutsheet.write(formatted_ap_info)

def add_wap_to_idf(switch_details):
    global legacy_switch_name
    global new_switch_name_list

    for switch in range(len(switch_details)):
        with open(f"{new_switch_name_list[switch-1]} switchport config.txt","a") as idf_config:
            with open(f"wap cutsheet {legacy_switch_name}.csv","r") as wap_cutsheet:
                wap_cutsheet.readline()

                for ap_details in wap_cutsheet:
                    list_of_ap_details = ap_details.split(",")

                    if list_of_ap_details[1] == new_switch_name_list[switch-1]:
                        interface = list_of_ap_details[2]
                        port_description = list_of_ap_details[7]
                        ap_vlan = list_of_ap_details[6]

                        idf_config.write(f"interface {interface}\ndescription {port_description}\nswitchport access {ap_vlan}\n!\n")

def excel_workbook_maker():
    global legacy_switch_name

    cutsheet_files = [f'{legacy_switch_name} cutsheet.csv', f'wap cutsheet {legacy_switch_name}.csv']

    wb = xl.Workbook()
    del wb[wb.sheetnames[0]]
    
    for file in cutsheet_files:

        with open(file) as f_name:
            if f'wap cutsheet {legacy_switch_name}.csv' == file:
                ws = wb.create_sheet(title='wap cutsheet')

                for row in csv.reader(f_name, delimiter=','):
                    ws.append(row)
                
            elif f'{legacy_switch_name} cutsheet.csv' == file:
                ws = wb.create_sheet(title=legacy_switch_name)

                for row in csv.reader(f_name, delimiter=','):
                    ws.append(row)

    wb.save(f'{legacy_switch_name} cutsheet.xlsx')

    os.remove(f'wap cutsheet {legacy_switch_name}.csv')
    os.remove(f'{legacy_switch_name} cutsheet.csv')
    os.remove(f'{legacy_switch_name} pre portmapped + patch_panel.csv')

